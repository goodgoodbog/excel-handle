# -*- coding=utf-8-*-
# 功能5：对多个订单与运单号进行合并
import os
import tkinter as tk
import tkinter.ttk as ttk
from tkinter import filedialog
import openpyxl
from common.all_use import choose_flie
from common.all_use import creat_thread
from common import all_use


class combined:
    def __init__(self, root):
        self.use = all_use.Use(root)
        self.six_Labeltip = tk.Label(root, text='功能5：合并多个订单与运单号:')
        self.six_Label1 = tk.Label(root, text='订单与运单号1:')
        self.six_Label2 = tk.Label(root, text='订单与运单号2:')
        self.six_Label3 = tk.Label(root, text='订单与运单号3:')
        self.six_Label4 = tk.Label(root, text='订单与运单号4:')
        self.six_Label5 = tk.Label(root, text='订单与运单号5:')
        self.six_Label6 = tk.Label(root, text='订单与运单号6:')
        self.six_Label7 = tk.Label(root, text='订单与运单号7:')
        self.six_Label8 = tk.Label(root, text='订单与运单号8:')
        self.six_Label9 = tk.Label(root, text='订单与运单号9:')
        self.six_Label_save = tk.Label(root, text='保存到:')
        self.six_Label_jg = tk.Label(root, text='结果提示:')
        self.six_Choose1 = ttk.Button(root, text='选择文件', command=lambda: choose_flie(self.six_Entry1))
        self.six_Choose2 = ttk.Button(root, text='选择文件', command=lambda: choose_flie(self.six_Entry2))
        self.six_Choose3 = ttk.Button(root, text='选择文件', command=lambda: choose_flie(self.six_Entry3))
        self.six_Choose4 = ttk.Button(root, text='选择文件', command=lambda: choose_flie(self.six_Entry4))
        self.six_Choose5 = ttk.Button(root, text='选择文件', command=lambda: choose_flie(self.six_Entry5))
        self.six_Choose6 = ttk.Button(root, text='选择文件', command=lambda: choose_flie(self.six_Entry6))
        self.six_Choose7 = ttk.Button(root, text='选择文件', command=lambda: choose_flie(self.six_Entry7))
        self.six_Choose8 = ttk.Button(root, text='选择文件', command=lambda: choose_flie(self.six_Entry8))
        self.six_Choose9 = ttk.Button(root, text='选择文件', command=lambda: choose_flie(self.six_Entry9))
        self.six_Choose_work = ttk.Button(root, text='开始合并', command=self.work)
        self.six_Button_choose_files = ttk.Button(root, text='选择多个表', command=self.choose_files)
        self.six_addChoose = ttk.Button(root, text='订单与运单号+1', command=self.add_Choose)
        self.six_Entry1 = tk.Entry(root, width=56, highlightthickness=1.5, highlightcolor='blue', name='six_Entry1',
                                   validate="key", relief='sunken',
                                   validatecommand=(self.use.check_open_file1, '%P', '%W'))
        self.six_Entry2 = tk.Entry(root, width=56, highlightthickness=1.5, highlightcolor='blue', name='six_Entry2',
                                   validate="key", relief='sunken',
                                   validatecommand=(self.use.check_open_file1, '%P', '%W'))
        self.six_Entry3 = tk.Entry(root, width=56, highlightthickness=1.5, highlightcolor='blue', name='six_Entry3',
                                   validate="key", relief='sunken',
                                   validatecommand=(self.use.check_open_file1, '%P', '%W'))
        self.six_Entry4 = tk.Entry(root, width=56, highlightthickness=1.5, highlightcolor='blue', name='six_Entry4',
                                   validate="key", relief='sunken',
                                   validatecommand=(self.use.check_open_file1, '%P', '%W'))
        self.six_Entry5 = tk.Entry(root, width=56, highlightthickness=1.5, highlightcolor='blue', name='six_Entry5',
                                   validate="key", relief='sunken',
                                   validatecommand=(self.use.check_open_file1, '%P', '%W'))
        self.six_Entry6 = tk.Entry(root, width=56, highlightthickness=1.5, highlightcolor='blue', name='six_Entry6',
                                   validate="key", relief='sunken',
                                   validatecommand=(self.use.check_open_file1, '%P', '%W'))
        self.six_Entry7 = tk.Entry(root, width=56, highlightthickness=1.5, highlightcolor='blue', name='six_Entry7',
                                   validate="key", relief='sunken',
                                   validatecommand=(self.use.check_open_file1, '%P', '%W'))
        self.six_Entry8 = tk.Entry(root, width=56, highlightthickness=1.5, highlightcolor='blue', name='six_Entry8',
                                   validate="key", relief='sunken',
                                   validatecommand=(self.use.check_open_file1, '%P', '%W'))
        self.six_Entry9 = tk.Entry(root, width=56, highlightthickness=1.5, highlightcolor='blue', name='six_Entry9',
                                   validate="key", relief='sunken',
                                   validatecommand=(self.use.check_open_file1, '%P', '%W'))
        self.six_Entry_save = tk.Entry(root, width=56, highlightthickness=1.5, highlightcolor='blue')
        self.six_Entry_jg = tk.Entry(root, width=56)

    #布局函数
    # @grid_progressbar
    def show(self):
        global b_list
        while len(all_use.tk_list):
            t = all_use.tk_list.pop()
            if t.widgetName == 'entry':
                t.delete("0", 'end')
            t.grid_forget()
        self.six_Labeltip.grid(row=2, column=0, sticky='w', columnspan=2)
        all_use.tk_list.append(self.six_Labeltip)

        self.six_Label1.grid(row=3, column=0, sticky='e')
        all_use.tk_list.append(self.six_Label1)

        self.six_Entry1.delete("0", 'end')
        self.six_Entry1.grid(row=3, column=1, sticky='e')
        all_use.tk_list.append(self.six_Entry1)
        self.six_Choose1.grid(row=3, column=2, sticky='w')
        all_use.tk_list.append(self.six_Choose1)

        self.six_Button_choose_files.grid(row=3, column=3, sticky='w')
        all_use.tk_list.append(self.six_Button_choose_files)

        self.six_addChoose.grid(row=3, column=4, sticky='w')
        all_use.tk_list.append(self.six_addChoose)

        self.six_Label_save.grid(row=12, column=0, sticky='e')
        all_use.tk_list.append(self.six_Label_save)
        self.six_Entry_save.delete("0", 'end')
        self.six_Entry_save.grid(row=12, column=1, sticky='w')
        all_use.tk_list.append(self.six_Entry_save)

        self.six_Choose_work.grid(row=13, column=1)
        all_use.tk_list.append(self.six_Choose_work)

        self.six_Label_jg.grid(row=14, column=0, sticky='e')
        all_use.tk_list.append(self.six_Label_jg)
        self.six_Entry_jg.delete("0", 'end')
        self.six_Entry_jg.grid(row=14, column=1, sticky='w')
        all_use.tk_list.append(self.six_Entry_jg)
        b_list = 8

    # TODO:还未转为polars
    # 数据处理函数
    @creat_thread
    def work(self):
        wb = openpyxl.Workbook()
        sheet1 = wb.active
        if self.six_Entry1.get() is not None:
            self.adds(sheet1, self.six_Entry1.get(), '订单与运单号1')
        if self.six_Entry2.get() is not None:
            self.adds(sheet1, self.six_Entry2.get(), '订单与运单号2')
        if self.six_Entry3.get() is not None:
            self.adds(sheet1, self.six_Entry3.get(), '订单与运单号3')
        if self.six_Entry4.get() is not None:
            self.adds(sheet1, self.six_Entry4.get(), '订单与运单号4')
        if self.six_Entry5.get() is not None:
            self.adds(sheet1, self.six_Entry5.get(), '订单与运单号5')
        if self.six_Entry6.get() is not None:
            self.adds(sheet1, self.six_Entry6.get(), '订单与运单号6')
        if self.six_Entry7.get() is not None:
            self.adds(sheet1, self.six_Entry7.get(), '订单与运单号7')
        if self.six_Entry8.get() is not None:
            self.adds(sheet1, self.six_Entry8.get(), '订单与运单号8')
        if self.six_Entry9.get() is not None:
            self.adds(sheet1, self.six_Entry9.get(), '订单与运单号9')

        s = os.path.dirname(self.six_Entry1.get())
        s1 = [s, self.six_Entry_save.get()]
        s2 = '\\'.join(s1)
        s3 = [s2, '.xlsx']
        s4 = ''.join(s3)
        try:
            wb.save(s4)
        except:
            self.six_Entry_jg.delete("0", 'end')
            self.six_Entry_jg.insert(tk.INSERT, '文件保存失败')
            self.six_Entry_jg.configure(fg='red')
            return
        self.six_Entry_jg.delete("0", 'end')
        self.six_Entry_jg.insert(tk.INSERT, '保存到%s中' % s4)
        self.six_Entry_jg.configure(fg='green')

    def add_Choose(self):
        global b_list
        i = b_list
        if i == 8:
            self.six_Label2.grid(row=4, column=0, sticky='e')
            all_use.tk_list.append(self.six_Label2)
            self.six_Entry2.grid(row=4, column=1, sticky='w')
            all_use.tk_list.append(self.six_Entry2)
            self.six_Choose2.grid(row=4, column=2, sticky='w')
            all_use.tk_list.append(self.six_Choose2)
            b_list = b_list - 1
        elif i == 7:
            self.six_Label3.grid(row=5, column=0, sticky='e')
            all_use.tk_list.append(self.six_Label3)
            self.six_Entry3.grid(row=5, column=1, sticky='w')
            all_use.tk_list.append(self.six_Entry3)
            self.six_Choose3.grid(row=5, column=2, sticky='w')
            all_use.tk_list.append(self.six_Choose3)
            b_list = b_list - 1
        elif i == 6:
            self.six_Label4.grid(row=6, column=0, sticky='e')
            all_use.tk_list.append(self.six_Label4)
            self.six_Entry4.grid(row=6, column=1, sticky='w')
            all_use.tk_list.append(self.six_Entry4)
            self.six_Choose4.grid(row=6, column=2, sticky='w')
            all_use.tk_list.append(self.six_Choose4)
            b_list = b_list - 1
        elif i == 5:
            self.six_Label5.grid(row=7, column=0, sticky='e')
            all_use.tk_list.append(self.six_Label5)
            self.six_Entry5.grid(row=7, column=1, sticky='w')
            all_use.tk_list.append(self.six_Entry5)
            self.six_Choose5.grid(row=7, column=2, sticky='w')
            all_use.tk_list.append(self.six_Choose5)
            b_list = b_list - 1
        elif i == 4:
            self.six_Label6.grid(row=8, column=0, sticky='e')
            all_use.tk_list.append(self.six_Label6)
            self.six_Entry6.grid(row=8, column=1, sticky='w')
            all_use.tk_list.append(self.six_Entry6)
            self.six_Choose6.grid(row=8, column=2, sticky='w')
            all_use.tk_list.append(self.six_Choose6)
            b_list = b_list - 1
        elif i == 3:
            self.six_Label7.grid(row=9, column=0, sticky='e')
            all_use.tk_list.append(self.six_Label7)
            self.six_Entry7.grid(row=9, column=1, sticky='w')
            all_use.tk_list.append(self.six_Entry7)
            self.six_Choose7.grid(row=9, column=2, sticky='w')
            all_use.tk_list.append(self.six_Choose7)
            b_list = b_list - 1
        elif i == 2:
            self.six_Label8.grid(row=10, column=0, sticky='e')
            all_use.tk_list.append(self.six_Label8)
            self.six_Entry8.grid(row=10, column=1, sticky='w')
            all_use.tk_list.append(self.six_Entry8)
            self.six_Choose8.grid(row=10, column=2, sticky='w')
            all_use.tk_list.append(self.six_Choose8)
            b_list = b_list - 1
        elif i == 1:
            self.six_Label9.grid(row=11, column=0, sticky='e')
            all_use.tk_list.append(self.six_Label9)
            self.six_Entry9.grid(row=11, column=1, sticky='w')
            all_use.tk_list.append(self.six_Entry9)
            self.six_Choose9.grid(row=11, column=2, sticky='w')
            all_use.tk_list.append(self.six_Choose9)
            b_list = b_list - 1

    ##一次性选择多个表
    def choose_files(self):
        global b_list
        file_paths = filedialog.askopenfilenames()
        if file_paths == '':
            return
        b_list = 8
        more_list = [[self.six_Label2, self.six_Entry2, self.six_Choose2],
                     [self.six_Label3, self.six_Entry3, self.six_Choose3],
                     [self.six_Label4, self.six_Entry4, self.six_Choose4],
                     [self.six_Label5, self.six_Entry5, self.six_Choose5],
                     [self.six_Label6, self.six_Entry6, self.six_Choose6],
                     [self.six_Label7, self.six_Entry7, self.six_Choose7],
                     [self.six_Label8, self.six_Entry8, self.six_Choose8],
                     [self.six_Label9, self.six_Entry9, self.six_Choose9]
                     ]

        more_list = iter(more_list)
        self.six_Entry1.delete("0", 'end')
        self.six_Entry1.insert(tk.INSERT, file_paths[0])
        now_list = next(more_list)
        list2 = [self.six_Entry2,
                 self.six_Entry3,
                 self.six_Entry4,
                 self.six_Entry5,
                 self.six_Entry6,
                 self.six_Entry7,
                 self.six_Entry8,
                 self.six_Entry9,
                 ]
        for file_path, entry in zip(file_paths[1:], list2):
            entry.delete("0", 'end')
            self.add_Choose()
            entry.insert(tk.INSERT, file_path)
            now_list = next(more_list)
        while now_list is not None:
            for data in now_list:
                if data.widgetName == 'entry':
                    data.delete("0", 'end')
                data.grid_forget()
            try:
                now_list = next(more_list)
            except StopIteration:
                break

    def adds(self, sheet1, filename, tip_filename):
        try:
            wb2 = openpyxl.load_workbook(filename)
        except:
            self.six_Entry_jg.delete("0", 'end')
            mmmm = tip_filename + '打开失败，请重新选择！！！'
            self.six_Entry_jg.insert(tk.INSERT, mmmm)
            self.six_Entry_jg.configure(fg='red')
            return
        sheet2 = wb2.active
        row1 = sheet1.max_row
        if sheet1.cell(1, 1).value is not None:
            sheet2.delete_rows(idx=1, amount=2, sticky='w')
        else:
            row1 = row1 - 1
        max_row2 = sheet2.max_row
        max_column2 = sheet2.max_column
        for i in range(1, max_row2 + 1):
            for j in range(1, max_column2 + 1):
                v = sheet2.cell(i, j).value
                if v:
                    sheet1.cell(i + row1, j, v)
