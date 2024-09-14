# -*- coding=utf-8-*-
# 功能7：扫码前订单拼接
import os
import time
from tkinter import ttk
import tkinter as tk
import openpyxl
from common.all_use import choose_flie
from common.all_use import creat_thread
from common import all_use


class splice:
    def __init__(self, root):
        use = all_use.Use(root)
        self.eight_Label1 = tk.Label(root, text='要发货的订单表:')
        self.eight_Label2 = tk.Label(root, text='物流商导出运单表:')
        self.eight_Label_save = tk.Label(root, text='保存到:')
        self.eight_Label_jg = tk.Label(root, text='结果提示:')
        self.eight_Labeltip = tk.Label(root, text='功能7：扫码前订单拼接')
        self.eight_Choose1 = ttk.Button(root, text='选择文件', command=lambda: choose_flie(self.eight_Entry1))
        self.eight_Choose2 = ttk.Button(root, text='选择文件', command=lambda: choose_flie(self.eight_Entry2))
        self.eight_Choose_work = ttk.Button(root, text='开始统计', command=self.work)
        self.eight_Entry1 = tk.Entry(root, width=56, highlightthickness=1.5, highlightcolor='blue', name='eight_Entry1',
                                     validate="key", relief='sunken',
                                     validatecommand=(use.check_open_file1, '%P', '%W'))
        self.eight_Entry2 = tk.Entry(root, width=56, highlightthickness=1.5, highlightcolor='blue', name='eight_Entry2',
                                     validate="key", relief='sunken',
                                     validatecommand=(use.check_open_file1, '%P', '%W'))
        self.eight_Entry_save = tk.Entry(root, width=56, highlightthickness=1.5, highlightcolor='blue')
        self.eight_Entry_jg = tk.Entry(root, width=56)

    # 布局函数
    # @grid_progressbar
    def show(self):
        while len(all_use.tk_list):
            t = all_use.tk_list.pop()
            if t.widgetName == 'entry':
                t.delete("0", 'end')
            t.grid_forget()
        self.eight_Labeltip.grid(row=2, column=0, sticky='w', columnspan=2)
        all_use.tk_list.append(self.eight_Labeltip)
        self.eight_Label1.grid(row=3, column=0, sticky='e')
        all_use.tk_list.append(self.eight_Label1)
        self.eight_Entry1.delete("0", 'end')
        self.eight_Entry1.grid(row=3, column=1, sticky='w')
        all_use.tk_list.append(self.eight_Entry1)
        self.eight_Choose1.grid(row=3, column=2, sticky='w')
        all_use.tk_list.append(self.eight_Choose1)
        self.eight_Label2.grid(row=4, column=0, sticky='e')
        all_use.tk_list.append(self.eight_Label2)
        self.eight_Entry2.delete("0", 'end')
        self.eight_Entry2.grid(row=4, column=1, sticky='w')
        all_use.tk_list.append(self.eight_Entry2)
        self.eight_Choose2.grid(row=4, column=2, sticky='w')
        all_use.tk_list.append(self.eight_Choose2)

        self.eight_Label_save.grid(row=5, column=0, sticky='e')
        all_use.tk_list.append(self.eight_Label_save)
        self.eight_Entry_save.delete("0", 'end')
        self.eight_Entry_save.grid(row=5, column=1, sticky='w')
        all_use.tk_list.append(self.eight_Entry_save)
        self.eight_Choose_work.grid(row=6, column=1)
        all_use.tk_list.append(self.eight_Choose_work)
        self.eight_Label_jg.grid(row=7, column=0, sticky='e')
        all_use.tk_list.append(self.eight_Label_jg)
        self.eight_Entry_jg.delete("0", 'end')
        self.eight_Entry_jg.grid(row=7, column=1, sticky='w')
        all_use.tk_list.append(self.eight_Entry_jg)
        # 获取时间
        now_time = time.strftime('%M分%S秒', time.localtime(time.time()))
        # 显示生成的名称
        self.eight_Entry_save.delete("0", 'end')
        self.eight_Entry_save.insert(tk.INSERT, '扫码前订单拼接-' + now_time)

    # TODO：未polars
    # 数据处理函数
    @creat_thread
    def work(self):
        try:
            wb1 = openpyxl.load_workbook(self.eight_Entry1.get())
        except:
            self.eight_Entry_jg.delete("0", 'end')
            self.eight_Entry_jg.insert(tk.INSERT, '要发货的订单表出错,请重新选择')
            self.eight_Entry_jg.configure(fg='red')
            return
        try:
            wb2 = openpyxl.load_workbook(self.eight_Entry2.get())
        except:
            self.eight_Entry_jg.delete("0", 'end')
            self.eight_Entry_jg.insert(tk.INSERT, '物流商导出运单表出错,请重新选择')
            self.eight_Entry_jg.configure(fg='red')
            return
        sheet1 = wb1.active
        sheet2 = wb2.active
        row1 = sheet1.max_row
        row2 = sheet2.max_row
        sheet1.insert_cols(idx=2)
        for i in range(1, row1 + 1):
            sss = sheet1.cell(i, 1).value
            for j in range(1, row2 + 1):
                if sss == sheet2.cell(j, 1).value:
                    sheet1.cell(i, 2).value = sheet2.cell(j, 2).value
        s = os.path.dirname(self.eight_Entry1.get())
        s1 = [s, self.eight_Entry_save.get()]
        s2 = '\\'.join(s1)
        s3 = [s2, '.xlsx']
        s4 = ''.join(s3)
        try:
            wb1.save(s4)
        except:
            self.seven_Entry_jg.delete("0", 'end')
            self.seven_Entry_jg.insert(tk.INSERT, '文件保存失败')
            self.eight_Entry_jg.configure(fg='red')
            return
        self.eight_Entry_jg.delete("0", 'end')
        self.eight_Entry_jg.insert(tk.INSERT, '保存到%s中' % s4)
        self.eight_Entry_jg.configure(fg='green')
