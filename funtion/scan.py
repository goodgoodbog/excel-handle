# -*- coding=utf-8-*-
# 功能8:扫码
import os
import time
from tkinter import ttk
import tkinter as tk
import openpyxl
from openpyxl.styles import PatternFill
from common.all_use import choose_flie
from common.all_use import creat_thread
from common import all_use


class scan:
    def __init__(self, root):
        self.use = all_use.Use(root)
        self.nine_Label1 = tk.Label(root, text='扫码匹配表:')
        self.nine_Label2 = tk.Label(root, text='扫码读取数据:')
        self.nine_Label_save = tk.Label(root, text='保存到:')
        self.nine_Label_jg = tk.Label(root, text='结果提示:')
        self.nine_Labeltip = tk.Label(root, text='功能8：扫码')
        self.nine_Choose1 = ttk.Button(root, text='选择文件', command=lambda: choose_flie(self.nine_Entry1))
        self.nine_Choose_work = ttk.Button(root, text='开始扫码', command=self.work)
        self.nine_Entry1 = tk.Entry(root, width=56, highlightthickness=1.5, highlightcolor='blue', name='nine_Entry1',
                                    validate="key", relief='sunken',
                                    validatecommand=(self.use.check_open_file1, '%P', '%W'))
        self.nine_Entry2 = tk.Entry(root, width=56, highlightthickness=1.5, highlightcolor='blue', name='nine_Entry2',
                                    validate="key", relief='sunken',
                                    validatecommand=(self.use.check_open_file1, '%P', '%W'))
        self.nine_Entry_save = tk.Entry(root, width=56, highlightthickness=1.5, highlightcolor='blue')
        self.nine_Entry_jg = tk.Entry(root, width=56)

    # 布局函数
    # @grid_progressbar
    def show(self):
        while len(all_use.tk_list):
            t = all_use.tk_list.pop()
            if t.widgetName == 'entry':
                t.delete("0", 'end')
            t.grid_forget()
        self.nine_Labeltip.grid(row=2, column=0, sticky='w', columnspan=2)
        all_use.tk_list.append(self.nine_Labeltip)
        self.nine_Label1.grid(row=3, column=0, sticky='e')
        all_use.tk_list.append(self.nine_Label1)
        self.nine_Entry1.delete("0", 'end')
        self.nine_Entry1.grid(row=3, column=1, sticky='w')
        all_use.tk_list.append(self.nine_Entry1)
        self.nine_Choose1.grid(row=3, column=2, sticky='w')
        all_use.tk_list.append(self.nine_Choose1)
        self.nine_Label2.grid(row=4, column=0, sticky='e')
        all_use.tk_list.append(self.nine_Label2)
        self.nine_Entry2.delete("0", 'end')
        self.nine_Entry2.grid(row=4, column=1, sticky='w')
        all_use.tk_list.append(self.nine_Entry2)

        self.nine_Label_save.grid(row=5, column=0, sticky='e')
        all_use.tk_list.append(self.nine_Label_save)
        self.nine_Entry_save.delete("0", 'end')
        self.nine_Entry_save.grid(row=5, column=1, sticky='w')
        all_use.tk_list.append(self.nine_Entry_save)
        self.nine_Choose_work.grid(row=6, column=1)
        all_use.tk_list.append(self.nine_Choose_work)
        self.nine_Label_jg.grid(row=7, column=0, sticky='e')
        all_use.tk_list.append(self.nine_Label_jg)
        self.nine_Entry_jg.delete("0", 'end')
        self.nine_Entry_jg.grid(row=7, column=1, sticky='w')
        all_use.tk_list.append(self.nine_Entry_jg)
        # 获取时间
        now_time = time.strftime('%M分%S秒', time.localtime(time.time()))
        # 显示生成的名称
        self.nine_Entry_save.delete("0", 'end')
        self.nine_Entry_save.insert(tk.INSERT, '扫码-' + now_time)

    # TODO：未polars
    @creat_thread
    def work(self):
        s1 = self.nine_Entry2.get()
        try:
            wb1 = openpyxl.load_workbook(self.nine_Entry1.get())
        except:
            self.nine_Entry_jg.delete("0", 'end')
            self.nine_Entry_jg.insert(tk.INSERT, '扫码匹配表出错,请重新选择')
            self.nine_Entry_jg.configure(fg='red')
            return
        sheet1 = wb1.active
        row1 = sheet1.max_row
        jg = 0
        red_font = PatternFill('solid', fgColor="FF0000")
        for i in range(1, row1 + 1):
            if sheet1.cell(i, 2).value == s1:
                jg = 1
                self.nine_Entry_jg.delete("0", 'end')
                self.nine_Entry_jg.insert(tk.INSERT, s1 + '成功')
                self.nine_Entry_jg.configure(fg='green')
                # 变颜色
                sheet1.cell(i, 2).fill = red_font
        if jg == 0:
            self.nine_Entry_jg.delete("0", 'end')
            self.nine_Entry_jg.insert('警告' + tk.INSERT, s1 + '不存在')
            self.nine_Entry_jg.configure(fg='red')

        s = os.path.dirname(self.nine_Entry1.get())
        s1 = [s, self.nine_Entry_save.get()]
        s2 = '\\'.join(s1)
        s3 = [s2, '.xlsx']
        s4 = ''.join(s3)
        try:
            wb1.save(s4)
        except:
            self.nine_Entry_jg.delete("0", 'end')
            self.nine_Entry_jg.insert(tk.INSERT, '文件保存失败')
            self.nine_Entry_jg.configure(fg='red')
            return
        self.nine_Entry2.delete("0", 'end')
        self.nine_Entry1.delete("0", 'end')
        self.nine_Entry1.insert(tk.INSERT, s4)