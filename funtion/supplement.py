# -*- coding=utf-8-*-
# 功能3：回填运单号订单列表处理
import os
import time
import tkinter as tk
import tkinter.ttk as ttk
import tkinter.filedialog
from tkinter import scrolledtext
import openpyxl
from common.all_use import choose_flie
from common.all_use import creat_thread
from common import all_use


class Supplement:
    def __init__(self, root):
        self.use = all_use.Use(root)
        self.d1 = tk.Label(root, text='功能3：补充信息：')
        self.d2 = tk.Label(root, text='E键发运单:')
        self.e1 = tk.Entry(root, name='supplement_e1', width=56, highlightthickness=1.5, highlightcolor='blue',
                           validate="key",
                           relief='sunken',
                           validatecommand=(self.use.check_open_file1, '%P', '%W'))
        self.choose1 = ttk.Button(root, text='选择文件', command=lambda: choose_flie(self.e1))
        self.d3 = tk.Label(root, text='订单列表，店铺导出:')
        self.e2 = tk.Entry(root, width=56, highlightthickness=1.5, highlightcolor='blue', name='supplement_e2',
                           validate="key",
                           relief='sunken',
                           validatecommand=(self.use.check_open_file1, '%P', '%W'))
        self.four_choose = ttk.Button(root, text='选择文件', command=self.four_choose_flie9)
        self.d4 = tk.Label(root, text='需要回填的订单号:')
        self.e3 = tk.Entry(root, width=56, highlightthickness=1.5, highlightcolor='blue', name='supplement_e3',
                           validate="key",
                           relief='sunken',
                           validatecommand=(self.use.check_open_file1, '%P', '%W'))
        self.choose3 = ttk.Button(root, text='选择文件', command=lambda: choose_flie(self.e3))
        self.d5 = tk.Label(root, text='保存到:')
        self.e11 = tk.Entry(root, width=56, highlightthickness=1.5, highlightcolor='blue')  ##保存文件名
        self.dBtton = ttk.Button(root, text='开始补充信息', command=self.work)
        self.jg = tk.Label(root, text='结果提示：')
        self.e10 = tk.Entry(root, width=56)
        self.scr = scrolledtext.ScrolledText(root, width=100)

    # 布局函数
    # @grid_progressbar
    def show(self):
        while len(all_use.tk_list):
            t = all_use.tk_list.pop()
            if t.widgetName == 'entry':
                t.delete("0", 'end')
            t.grid_forget()
        self.d1.grid(row=2, column=0, sticky='w', columnspan=2)
        all_use.tk_list.append(self.d1)
        self.d2.grid(row=3, column=0, sticky='e')
        all_use.tk_list.append(self.d2)
        self.e1.delete("0", 'end')
        self.e1.grid(row=3, column=1, sticky='w')
        all_use.tk_list.append(self.e1)
        self.choose1.grid(row=3, column=2, sticky='w')
        all_use.tk_list.append(self.choose1)
        self.d3.grid(row=4, column=0, sticky='e')
        all_use.tk_list.append(self.d3)
        self.e2.delete("0", 'end')
        self.e2.grid(row=4, column=1, sticky='w')
        all_use.tk_list.append(self.e2)

        self.four_choose.grid(row=4, column=2, sticky='w')
        all_use.tk_list.append(self.four_choose)

        self.d4.grid(row=5, column=0, sticky='e')
        all_use.tk_list.append(self.d4)
        self.e3.delete("0", 'end')
        self.e3.grid(row=5, column=1, sticky='w')
        all_use.tk_list.append(self.e3)
        self.choose3.grid(row=5, column=2, sticky='w')
        all_use.tk_list.append(self.choose3)
        self.d5.grid(row=6, column=0, sticky='e')
        all_use.tk_list.append(self.d5)
        self.e11.delete("0", 'end')
        self.e11.grid(row=6, column=1, sticky='w')
        all_use.tk_list.append(self.e11)
        self.dBtton.grid(row=7, column=1)
        all_use.tk_list.append(self.dBtton)
        self.jg.grid(row=8, column=0, sticky='e')
        all_use.tk_list.append(self.jg)
        self.e10.delete("0", 'end')
        self.e10.grid(row=8, column=1, sticky='w')
        all_use.tk_list.append(self.e10)

    # 数据处理函数
    @creat_thread
    def work(self):
        list2 = []
        global alllist_4
        try:
            wb1 = openpyxl.load_workbook(self.e1.get())
        except:
            self.e10.delete("0", 'end')
            self.e10.insert(tk.INSERT, 'E键发运单选择出错,请重新选择')
            self.e10.configure(fg='red')
            return
        try:
            wb2 = openpyxl.load_workbook(self.e2.get())
        except:
            self.e10.delete("0", 'end')
            self.e10.insert(tk.INSERT, '订单列表选择出错,请重新选择')
            self.e10.configure(fg='red')
            return
        try:
            wb3 = openpyxl.load_workbook(self.e3.get())
        except:
            self.e10.delete("0", 'end')
            self.e10.insert(tk.INSERT, '需要回填的订单号选择出错,请重新选择')
            self.e10.configure(fg='red')
            return
        sheet1 = wb1.active
        sheet1.delete_cols(idx=1, amount=1)  ##删除第一列
        sheet2 = wb2.active
        sheet3 = wb3.active
        row1 = sheet1.max_row
        row2 = sheet2.max_row
        row3 = sheet3.max_row
        count = 0
        for i in range(1, row3 + 1):
            if sheet3.cell(i, 1).value is not None:
                alllist_4.append(sheet3.cell(i, 1).value)
        try:
            alllist_4 = list(set(alllist_4))
        except:
            pass
        # 寻找订单编号
        column2 = sheet2.max_column

        Order_Numbe2 = -1
        fa_huo_shu_liang2 = -1
        sheng_yu_shu_liang2 = -1
        wu_liu_shang2 = -1
        wu_liu_dan_hao2 = -1
        for i in range(1, column2 + 1):
            if sheet2.cell(2, i).value == '订单编号':
                Order_Numbe2 = i
                continue
            if sheet2.cell(2, i).value == '发货数量':
                fa_huo_shu_liang2 = i
                continue
            if sheet2.cell(2, i).value == '剩余可发货数量':
                sheng_yu_shu_liang2 = i
                continue
            if sheet2.cell(2, i).value == '物流商':
                wu_liu_shang2 = i
                continue
            if sheet2.cell(2, i).value == '物流单号':
                wu_liu_dan_hao2 = i
                continue
        # 寻找订单号
        column1 = sheet1.max_column
        Order_Numbe1 = -1
        for i in range(1, column1 + 1):
            if sheet1.cell(1, i).value == '订单号':
                Order_Numbe1 = i
                break
        # 寻找运单号
        yun_dan_hao1 = -1
        for i in range(1, column1 + 1):
            if sheet1.cell(1, i).value == '运单号':
                yun_dan_hao1 = i
                break
        # 检查文件中是否存在某列
        if Order_Numbe2 == -1:
            self.e10.delete("0", 'end')
            self.e10.insert(tk.INSERT, "'订单编号'不存在，请检查文件是否选错！！！")
            self.e2.configure(fg='orange')
            return
        elif fa_huo_shu_liang2 == -1:
            self.e10.delete("0", 'end')
            self.e10.insert(tk.INSERT, "'发货数量'不存在，请检查文件是否选错！！！")
            self.e2.configure(fg='orange')
            return
        elif sheng_yu_shu_liang2 == -1:
            self.e10.delete("0", 'end')
            self.e10.insert(tk.INSERT, "'剩余可发货数量'不存在，请检查文件是否选错！！！")
            self.e2.configure(fg='orange')
            return
        elif wu_liu_shang2 == -1:
            self.e10.delete("0", 'end')
            self.e10.insert(tk.INSERT, "'物流商'不存在，请检查文件是否选错！！！")
            self.e2.configure(fg='orange')
            return
        elif wu_liu_dan_hao2 == -1:
            self.e10.delete("0", 'end')
            self.e10.insert(tk.INSERT, "'物流单号'不存在，请检查文件是否选错！！！")
            self.e2.configure(fg='orange')
            return
        elif Order_Numbe1 == -1:
            self.e10.delete("0", 'end')
            self.e10.insert(tk.INSERT, "'订单号'不存在，请检查文件是否选错！！！")
            self.e1.configure(fg='orange')
            return
        elif yun_dan_hao1 == -1:
            self.e10.delete("0", 'end')
            self.e10.insert(tk.INSERT, "'运单号'不存在，请检查文件是否选错！！！")
            self.e1.configure(fg='orange')
            return

        for i in range(3, row2 + 1):
            s = sheet2.cell(i, Order_Numbe2).value
            if s[0] in list2:
                pass
            else:
                list2.append(s[0])
            for j in range(2, row1 + 1):
                if s == sheet1.cell(j, Order_Numbe1).value:
                    for k in range(1, row3 + 1):
                        if s == sheet3.cell(k, 1).value:
                            sheet2.cell(i, wu_liu_dan_hao2).value = sheet1.cell(j, yun_dan_hao1).value
                            sheet2.cell(i, wu_liu_shang2).value = 'yanwen'
                            sheet2.cell(i, fa_huo_shu_liang2).value = sheet2.cell(i, sheng_yu_shu_liang2).value
                            try:
                                alllist_4.remove(s)
                                count = count + 1
                            except:
                                continue
        self.scr.grid(row=99, columnspan=99, sticky='w')
        all_use.tk_list.append(self.scr)
        self.scr.delete("1.0", 'end')
        self.scr.insert("end", "要回填的订单表中，没有匹配到的订单号：\n")
        for i in alllist_4:
            if i[0] in list2:
                self.scr.insert("end", str(i) + "\n")
        s = os.path.dirname(self.e1.get())
        s1 = [s, self.e11.get()]
        s2 = '\\'.join(s1)
        s3 = [s2, '.xlsx']
        s4 = ''.join(s3)
        try:
            wb2.save(s4)
        except:
            self.e10.delete("0", 'end')
            self.e10.insert(tk.INSERT, '文件保存失败')
            self.e10.configure(fg='red')
            return count
        q = str(count)
        self.e10.delete("0", 'end')
        self.e10.insert(tk.INSERT, '订单修改数为' + q + '，保存到文件：' + s4)
        self.e10.configure(fg='green')

    def four_choose_flie9(self):
        need_ = 0
        fileName = tkinter.filedialog.askopenfilename()

        # 自动生成名称
        for i in range(0, len(fileName)):
            if fileName[-i] == "/":
                need_ = 0 - i
                break
        save_name = fileName[-6:need_:-1]
        save_name = save_name[::-1]
        # 获取时间
        now_time = time.strftime('%M分%S秒', time.localtime(time.time()))
        # 显示生成的名称
        save_name = save_name + '-' + now_time
        self.e11.delete("0", 'end')
        self.e11.insert(tk.INSERT, save_name)

        fileName = fileName.replace("/", "\\")
        if '\\' in fileName:
            self.e2.delete("0", 'end')
            self.e2.insert(tk.INSERT, fileName)
